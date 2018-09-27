# _*_ coding: utf -8 _*_
import xlrd, xlwt
from xlutils.copy import copy
import xlutils
import xlsxwriter
import openpyxl
from openpyxl.cell.cell import WriteOnlyCell

class Update_Excel(object):

    def update_hkcost_sysbill(self, file, value):
        self.data = xlrd.open_workbook(file)
        self.table = self.data.sheet_by_index(0)
        self.d = xlutils.copy.copy(self.data)
        w = self.d.get_sheet(0)
        w.write_string(1, 1, value)
        w.write(2, 1, value)
        self.d.save(file)
    def update_hkcost_sysbill2(self,file,value,outputfile):
        self.date = xlrd.open_workbook(file)
        self.table = self.date.sheet_by_index(0)
        self.nrows = self.table.nrows
        self.ncols = self.table.ncols
        ctype = 2
        xf = 0
        self.table.put_cell(1, 1, ctype, value, xf)
        self.table.put_cell(2, 1, ctype, value, xf)
        #
        self.wbook = xlwt.Workbook()
        self.wsheet = self.wbook.add_sheet(self.table.name, cell_overwrite_ok = True)
        style = xlwt.XFStyle()
        for r in range(self.nrows):
            for c in range(self.ncols):
                fnt = xlwt.Font()
                fnt.name = u'微软雅黑'
                fnt.bold = False
                style.font = fnt
                self.wsheet.write_string(r, c, self.table.cell_value(r, c), style)
                # self.wbook.set_cell_text(r, c, self.table.cell_value(r, c), style)
        self.wbook.save(outputfile)

    def update_hkcost_sysbill3(self, file, value, outputfile):
        self.date = xlrd.open_workbook(file)
        self.table = self.date.sheet_by_index(0)
        self.nrows = self.table.nrows
        self.ncols = self.table.ncols
        ctype = 2
        xf = 0
        self.table.put_cell(1, 1, ctype, value, xf)
        self.table.put_cell(2, 1, ctype, value, xf)
        self.workbook = xlsxwriter.Workbook(outputfile)
        self.worksheet = self.workbook.add_worksheet('账单信息')
        for r in range(self.nrows):
            for c in range(self.ncols):
                self.worksheet.write_string(r, c, self.table.cell_value(r, c))
        self.workbook.close()
    def update_hkcost_sysbill4(self, file, value, outputfile):
        self.date = xlrd.open_workbook(file)
        self.table = self.date.sheet_by_index(0)
        self.nrows = self.table.nrows
        self.ncols = self.table.ncols
        ctype = 2
        xf = 0
        self.table.put_cell(1, 1, ctype, value, xf)
        self.table.put_cell(2, 1, ctype, value, xf)


        # self.wb = openpyxl.Workbook(write_only=True)
        # self.sh = self.wb.create_sheet('账单信息')
        # self.cell = WriteOnlyCell(self.wb)
        # for r in range(self.nrows):
        #     for c in range(self.ncols):
        #         self.cell.value =  self.table.cell_value(r, c)
        # self.wb.save(outputfile)

        # self.wb1 = openpyxl.load_workbook(file)
        # self.old_sheet = self.wb1.get_sheet_by_name('账单信息')
        # self.old_sheet.title = 'template'
        # self.wb1.creat_sheet('账单', 1)
        # new_sheet = wb1.get_sheet_by_name('账单')
        # celStyle = self.NamedStyle(name='celStyle')




# uhs = Update_Excel()
# uhs.update_hkcost_sysbill('账单信息导入模板.xls', '5001013')
# uhs.update_hkcost_sysbill2('账单信息导入模板.xls', '5001013', 'output.xls')
# uhs.update_hkcost_sysbill3('账单信息导入模板.xls','5001025','output.xlsx')
# uhs.update_hkcost_sysbill4('账单信息导入模板.xls','5001013','output.xlsx')